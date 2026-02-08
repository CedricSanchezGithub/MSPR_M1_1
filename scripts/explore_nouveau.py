#!/usr/bin/env python3
"""
Script d'exploration légère des nouveaux fichiers dans data/input/nouveau/
Lit uniquement les métadonnées et les 30 premières lignes pour éviter les problèmes mémoire.
"""

import os
import sys

try:
    import pandas as pd
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

INPUT_DIR = "data/input/nouveau"
OUTPUT_FILE = "outputs/exploration_nouveau_output.txt"
MAX_LIGNES = 30


def log(msg, output_lines):
    print(msg)
    output_lines.append(msg)


def explorer_xlsx(filepath, output_lines):
    """Explore un fichier xlsx sans tout charger en mémoire."""
    filename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath) / (1024 * 1024)

    log(f"\n{'=' * 70}", output_lines)
    log(f"FICHIER: {filename} ({filesize:.1f} MB)", output_lines)
    log("=" * 70, output_lines)

    # Ouvrir avec openpyxl en read_only pour ne pas tout charger
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    log(f"\n  Nombre d'onglets: {len(sheet_names)}", output_lines)
    log(f"  Onglets: {', '.join(sheet_names)}", output_lines)

    for sheet_name in sheet_names:
        log(f"\n  {'─' * 60}", output_lines)
        log(f"  ONGLET: {sheet_name}", output_lines)
        log(f"  {'─' * 60}", output_lines)

        # Lire seulement les N premières lignes avec pandas
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=MAX_LIGNES, engine='openpyxl')
        except Exception as e:
            log(f"    ERREUR de lecture: {e}", output_lines)
            continue

        log(f"  Colonnes ({len(df.columns)}):", output_lines)
        for i, col in enumerate(df.columns, 1):
            dtype = df[col].dtype
            non_null = df[col].notna().sum()
            log(f"    {i:3}. {str(col):<40} | type: {str(dtype):<10} | non-null: {non_null}/{len(df)}", output_lines)

        # Compter le nombre total de lignes via openpyxl (sans tout charger)
        ws = wb[sheet_name]
        total_rows = ws.max_row - 1 if ws.max_row else 0  # -1 pour le header
        log(f"\n  Lignes totales (estimé): {total_rows:,}", output_lines)

        # Aperçu des 5 premières lignes
        log(f"\n  APERÇU (5 premières lignes):", output_lines)
        sample = df.head(5).to_string(max_colwidth=50)
        for line in sample.split('\n'):
            log(f"    {line}", output_lines)

        # Stats basiques sur les colonnes numériques
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        if numeric_cols:
            log(f"\n  STATS (colonnes numériques, sur {MAX_LIGNES} lignes):", output_lines)
            stats = df[numeric_cols].describe().round(2)
            for line in stats.to_string().split('\n'):
                log(f"    {line}", output_lines)

        # Valeurs uniques pour colonnes catégorielles (peu de valeurs distinctes)
        cat_cols = df.select_dtypes(include='object').columns.tolist()
        if cat_cols:
            log(f"\n  VALEURS UNIQUES (colonnes texte):", output_lines)
            for col in cat_cols:
                n_unique = df[col].nunique()
                if n_unique <= 20:
                    vals = df[col].value_counts().head(10)
                    log(f"    {col} ({n_unique} valeurs):", output_lines)
                    for val, count in vals.items():
                        log(f"      - {str(val)[:50]}: {count}", output_lines)
                else:
                    log(f"    {col}: {n_unique} valeurs uniques (trop pour lister)", output_lines)

    wb.close()


def main():
    output_lines = []

    log("=" * 70, output_lines)
    log("ANALYSE EXPLORATOIRE - NOUVEAUX FICHIERS (data/input/nouveau/)", output_lines)
    log("=" * 70, output_lines)

    # Lister les fichiers
    files = [f for f in os.listdir(INPUT_DIR)
             if f.endswith(('.xlsx', '.xls', '.csv', '.CSV', '.txt'))]

    if not files:
        log(f"\nAucun fichier trouvé dans {INPUT_DIR}/", output_lines)
        return

    log(f"\nFichiers trouvés:", output_lines)
    for f in sorted(files):
        size = os.path.getsize(os.path.join(INPUT_DIR, f)) / (1024 * 1024)
        log(f"  - {f} ({size:.1f} MB)", output_lines)

    # Analyser chaque fichier
    for f in sorted(files):
        filepath = os.path.join(INPUT_DIR, f)
        if f.endswith(('.xlsx', '.xls')):
            explorer_xlsx(filepath, output_lines)
        elif f.endswith(('.csv', '.CSV', '.txt')):
            log(f"\n  {f}: fichier CSV, utiliser explore_diplomes.py", output_lines)

    log(f"\n{'=' * 70}", output_lines)
    log("FIN DE L'ANALYSE", output_lines)
    log("=" * 70, output_lines)

    # Sauvegarder
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))

    print(f"\n=> Résultats sauvegardés dans: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
