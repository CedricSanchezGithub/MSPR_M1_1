#!/usr/bin/env python3
"""
Exploration légère des fichiers xlsx dans data/input/nouveau/
Détecte le vrai header INSEE (lignes de métadonnées à sauter).
"""

import os
import sys

try:
    import pandas as pd
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd

import openpyxl

INPUT_DIR = "data/input/nouveau"
OUTPUT_DIR = "outputs"
MAX_LIGNES = 30


def log(msg, output_lines):
    print(msg)
    output_lines.append(msg)


def explore_sheet(filepath, sheet_name, output_lines):
    """Explore un onglet d'un fichier xlsx."""
    log(f"\n  {'─' * 60}", output_lines)
    log(f"  ONGLET: {sheet_name}", output_lines)
    log(f"  {'─' * 60}", output_lines)

    # Lire brut pour détecter le header
    df_raw = pd.read_excel(filepath, sheet_name=sheet_name, nrows=15, header=None, engine='openpyxl')

    # Afficher les premières lignes brutes
    log(f"\n  Lignes brutes (détection header):", output_lines)
    for idx, row in df_raw.head(12).iterrows():
        vals = [str(v)[:35] for v in row.values[:6]]
        log(f"    Ligne {idx}: {vals}", output_lines)

    # Chercher le header : ligne avec CODGEO ou des noms de variables INSEE
    header_row = None
    for idx, row in df_raw.iterrows():
        vals_str = ' '.join([str(v).upper() for v in row.values if pd.notna(v)])
        if any(k in vals_str for k in ['CODGEO', 'CODE GEO', 'LIBGEO', 'LIBELLE']):
            header_row = idx
            break

    if header_row is not None:
        log(f"\n  Header détecté en ligne {header_row}", output_lines)
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row, nrows=MAX_LIGNES, engine='openpyxl')
    else:
        log(f"\n  Header non détecté, ligne 0 utilisée", output_lines)
        df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=MAX_LIGNES, engine='openpyxl')

    # Nombre de lignes total
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    total_rows = ws.max_row - (header_row + 1 if header_row is not None else 1)
    wb.close()

    log(f"  Lignes estimées: {total_rows:,}", output_lines)
    log(f"  Colonnes: {len(df.columns)}", output_lines)

    # Lister les colonnes
    log(f"\n  COLONNES ({len(df.columns)}):", output_lines)
    for i, col in enumerate(df.columns, 1):
        dtype = str(df[col].dtype)
        non_null = df[col].notna().sum()
        log(f"    {i:3}. {str(col):<50} | type: {dtype:<10} | non-null: {non_null}/{len(df)}", output_lines)

    # Colonnes géographiques
    geo_cols = [c for c in df.columns if any(k in str(c).upper() for k in
        ['CODGEO', 'CODE', 'COMMUNE', 'DEP', 'REGION', 'GEO', 'INSEE', 'LIB'])]
    if geo_cols:
        log(f"\n  Colonnes géographiques: {geo_cols}", output_lines)

    # Détecter les années dans les noms de colonnes
    import re
    years = set()
    for col in df.columns:
        found = re.findall(r'(19[6-9]\d|20[0-2]\d)', str(col))
        years.update(found)
    if years:
        log(f"  Années détectées dans les colonnes: {sorted(years)}", output_lines)

    # Aperçu
    log(f"\n  APERÇU (5 premières lignes):", output_lines)
    sample = df.head(5).to_string(max_colwidth=35)
    for line in sample.split('\n'):
        log(f"    {line}", output_lines)

    # Stats numériques (juste min/max/mean)
    numeric_cols = df.select_dtypes(include='number').columns.tolist()
    if numeric_cols:
        log(f"\n  STATS RAPIDES ({len(numeric_cols)} colonnes numériques):", output_lines)
        for col in numeric_cols[:10]:
            vals = df[col].dropna()
            if len(vals) > 0:
                log(f"    {str(col):<50} | min: {vals.min():.0f} | max: {vals.max():.0f} | moy: {vals.mean():.1f}", output_lines)
        if len(numeric_cols) > 10:
            log(f"    ... et {len(numeric_cols) - 10} autres colonnes numériques", output_lines)


def explore_file(filepath, output_lines):
    """Explore un fichier xlsx complet."""
    filename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath) / (1024 * 1024)

    log(f"\n{'=' * 70}", output_lines)
    log(f"FICHIER: {filename} ({filesize:.1f} MB)", output_lines)
    log("=" * 70, output_lines)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    log(f"\n  Onglets ({len(sheet_names)}): {', '.join(sheet_names[:10])}", output_lines)
    if len(sheet_names) > 10:
        log(f"  ... et {len(sheet_names) - 10} autres", output_lines)
    wb.close()

    # Explorer seulement le premier onglet de données (pas les onglets meta/doc)
    for sn in sheet_names:
        if sn.lower() in ['documentation', 'source', 'sources', 'notes']:
            continue
        explore_sheet(filepath, sn, output_lines)
        break  # Un seul onglet principal suffit pour l'exploration


def main():
    output_lines = []

    log("=" * 70, output_lines)
    log("ANALYSE EXPLORATOIRE - NOUVEAUX FICHIERS INSEE", output_lines)
    log("=" * 70, output_lines)

    files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(('.xlsx', '.xls'))])

    log(f"\nFichiers trouvés ({len(files)}):", output_lines)
    for f in files:
        size = os.path.getsize(os.path.join(INPUT_DIR, f)) / (1024 * 1024)
        log(f"  - {f} ({size:.1f} MB)", output_lines)

    for f in files:
        explore_file(os.path.join(INPUT_DIR, f), output_lines)

    log(f"\n{'=' * 70}", output_lines)
    log("FIN DE L'ANALYSE", output_lines)
    log("=" * 70, output_lines)

    output_file = os.path.join(OUTPUT_DIR, "exploration_nouveau_insee_output.txt")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))

    print(f"\n=> Résultats sauvegardés dans: {output_file}")


if __name__ == "__main__":
    main()
