"""Exploration du fichier arrêtés de catastrophes naturelles (GASPAR/CatNat)."""
import os

OUTPUT_FILE = "data/output/exploration_catnat.txt"
INPUT_FILE = "data/input/nouveau/Arretes_de_catastrophe_naturelles.xlsx"

os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

with open(OUTPUT_FILE, "w", encoding="utf-8") as out:
    size_mb = os.path.getsize(INPUT_FILE) / (1024 * 1024)
    out.write(f"FICHIER : {os.path.basename(INPUT_FILE)}\n")
    out.write(f"TAILLE  : {size_mb:.1f} MB\n\n")

    # Lire avec openpyxl en read_only pour les métadonnées
    from openpyxl import load_workbook
    wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    out.write(f"Onglets : {wb.sheetnames}\n\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n{'='*60}\n")
        out.write(f"ONGLET : {sheet_name}\n")
        out.write(f"{'='*60}\n\n")

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 30:
                break
            rows.append(row)

        if not rows:
            out.write("Onglet vide\n")
            continue

        # Trouver le header (première ligne non vide avec plusieurs valeurs)
        header_idx = 0
        for idx, row in enumerate(rows):
            non_none = sum(1 for v in row if v is not None)
            if non_none >= 3:
                header_idx = idx
                break

        header = [str(v) if v is not None else "" for v in rows[header_idx]]
        out.write(f"Header (ligne {header_idx}) : {header}\n")
        out.write(f"Nombre de colonnes : {len(header)}\n\n")

        out.write("--- 30 premières lignes ---\n")
        for i, row in enumerate(rows):
            vals = [str(v) if v is not None else "" for v in row[:8]]
            line = " | ".join(vals)
            if len(row) > 8:
                line += f" ... (+{len(row)-8} cols)"
            out.write(f"L{i:>3} | {line}\n")

    wb.close()

    # Maintenant lire avec pandas pour l'analyse
    import pandas as pd

    out.write(f"\n\n{'='*60}\n")
    out.write("ANALYSE PANDAS\n")
    out.write(f"{'='*60}\n\n")

    # Essayer de lire le premier onglet
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name=0)
        out.write(f"Shape : {df.shape}\n")
        out.write(f"Colonnes : {list(df.columns)}\n\n")

        out.write("--- dtypes ---\n")
        for col in df.columns:
            out.write(f"  {col} : {str(df[col].dtype)}\n")

        out.write("\n--- Valeurs uniques par colonne ---\n")
        for col in df.columns:
            nunique = df[col].nunique()
            out.write(f"\n  {col} ({nunique} valeurs uniques) :\n")
            if nunique <= 30:
                out.write(f"    Toutes : {sorted(df[col].dropna().unique().tolist())}\n")
            else:
                sample = df[col].dropna().unique()[:15].tolist()
                out.write(f"    Exemples : {sample}\n")

        # Vérifier la présence de code INSEE
        for col in df.columns:
            col_lower = col.lower()
            if any(k in col_lower for k in ["insee", "codgeo", "commune", "code"]):
                out.write(f"\n  >> Colonne potentielle de jointure : {col}\n")
                out.write(f"     Nb valeurs uniques : {df[col].nunique()}\n")
                out.write(f"     Exemples : {df[col].dropna().unique()[:10].tolist()}\n")

        # Vérifier les années
        for col in df.columns:
            col_lower = col.lower()
            if any(k in col_lower for k in ["date", "année", "annee", "year", "debut", "fin"]):
                out.write(f"\n  >> Colonne temporelle : {col}\n")
                out.write(f"     Min : {df[col].min()}\n")
                out.write(f"     Max : {df[col].max()}\n")
                out.write(f"     Exemples : {df[col].dropna().unique()[:10].tolist()}\n")

    except Exception as e:
        out.write(f"Erreur pandas : {e}\n")

print(f"Exploration terminée -> {OUTPUT_FILE}")
