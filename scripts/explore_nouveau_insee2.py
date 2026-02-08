#!/usr/bin/env python3
"""
Exploration ciblée des fichiers INSEE séries historiques.
Lit un onglet DEP_xxxx pour comprendre la structure des données.
"""

import os
import sys
import re

try:
    import pandas as pd
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd

import openpyxl

INPUT_DIR = "data/input/nouveau"
OUTPUT_FILE = "outputs/exploration_nouveau_insee2_output.txt"
MAX_LIGNES = 30


def log(msg, lines):
    print(msg)
    lines.append(msg)


def explore_file(filepath, lines):
    filename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath) / (1024 * 1024)

    log(f"\n{'=' * 70}", lines)
    log(f"FICHIER: {filename} ({filesize:.1f} MB)", lines)
    log("=" * 70, lines)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    log(f"  Onglets ({len(sheet_names)}): {', '.join(sheet_names)}", lines)
    wb.close()

    # Trouver les onglets de données (DEP_xxxx ou COM_xxxx)
    data_sheets = [s for s in sheet_names if s.startswith(('DEP_', 'COM_'))]
    if not data_sheets:
        data_sheets = [s for s in sheet_names if s not in ['Présentation', 'Documentation', 'Modifications_territoriales']]

    log(f"  Onglets de données: {data_sheets}", lines)

    # Explorer le premier et le dernier pour voir la structure
    sheets_to_check = []
    if data_sheets:
        sheets_to_check.append(data_sheets[0])
        if len(data_sheets) > 1:
            sheets_to_check.append(data_sheets[-1])

    for sheet_name in sheets_to_check:
        log(f"\n  {'─' * 60}", lines)
        log(f"  ONGLET: {sheet_name}", lines)
        log(f"  {'─' * 60}", lines)

        # Lire brut
        df_raw = pd.read_excel(filepath, sheet_name=sheet_name, nrows=15, header=None, engine='openpyxl')

        log(f"\n  Lignes brutes:", lines)
        for idx, row in df_raw.head(12).iterrows():
            vals = [str(v)[:40] for v in row.values[:8]]
            log(f"    Ligne {idx}: {vals}", lines)

        # Trouver le header
        header_row = None
        for idx, row in df_raw.iterrows():
            vals_str = ' '.join([str(v).upper() for v in row.values if pd.notna(v)])
            if any(k in vals_str for k in ['CODGEO', 'CODE GÉO', 'LIBGEO', 'LIBELLÉ']):
                header_row = idx
                break

        if header_row is not None:
            log(f"\n  Header en ligne {header_row}", lines)
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row, nrows=MAX_LIGNES, engine='openpyxl')
        else:
            # Essayer ligne 0
            df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=MAX_LIGNES, engine='openpyxl')
            log(f"\n  Header par défaut (ligne 0)", lines)

        # Compter les lignes
        wb2 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb2[sheet_name]
        total = ws.max_row - (header_row + 1 if header_row else 1)
        wb2.close()

        log(f"  Lignes: {total:,}", lines)
        log(f"  Colonnes: {len(df.columns)}", lines)

        # Lister les colonnes
        log(f"\n  COLONNES ({len(df.columns)}):", lines)
        for i, col in enumerate(df.columns, 1):
            dtype = str(df[col].dtype)
            non_null = df[col].notna().sum()
            log(f"    {i:3}. {str(col):<55} | {dtype:<10} | {non_null}/{len(df)}", lines)

        # Colonnes géo
        geo_cols = [c for c in df.columns if any(k in str(c).upper() for k in ['CODGEO', 'CODE', 'LIB', 'DEP', 'REG'])]
        if geo_cols:
            log(f"\n  Colonnes géo: {geo_cols}", lines)

        # Années
        years = set()
        for col in df.columns:
            found = re.findall(r'((?:19[6-9]|20[0-2])\d)', str(col))
            years.update(found)
        if years:
            log(f"  Années dans colonnes: {sorted(years)}", lines)

        # Aperçu
        log(f"\n  APERÇU (3 premières lignes):", lines)
        sample = df.head(3).to_string(max_colwidth=30)
        for line in sample.split('\n'):
            log(f"    {line}", lines)


def main():
    lines = []

    log("=" * 70, lines)
    log("ANALYSE EXPLORATOIRE - FICHIERS INSEE SÉRIES HISTORIQUES", lines)
    log("=" * 70, lines)

    files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(('.xlsx', '.xls'))])
    for f in files:
        size = os.path.getsize(os.path.join(INPUT_DIR, f)) / (1024 * 1024)
        log(f"  - {f} ({size:.1f} MB)", lines)

    for f in files:
        explore_file(os.path.join(INPUT_DIR, f), lines)

    log(f"\n{'=' * 70}", lines)
    log("FIN", lines)
    log("=" * 70, lines)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as out:
        out.write('\n'.join(lines))
    print(f"\n=> Sauvegardé dans: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
