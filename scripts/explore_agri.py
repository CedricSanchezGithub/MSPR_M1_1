#!/usr/bin/env python3
"""
Exploration légère du fichier agri.xlsx
"""

import os
import sys

try:
    import pandas as pd
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd

import openpyxl

INPUT_FILE = "data/input/nouveau/agri.xlsx"
OUTPUT_FILE = "outputs/exploration_agri_output.txt"
MAX_LIGNES = 30


def log(msg, output_lines):
    print(msg)
    output_lines.append(msg)


def main():
    output_lines = []

    filename = os.path.basename(INPUT_FILE)
    filesize = os.path.getsize(INPUT_FILE) / (1024 * 1024)

    log("=" * 70, output_lines)
    log(f"ANALYSE EXPLORATOIRE - {filename} ({filesize:.1f} MB)", output_lines)
    log("=" * 70, output_lines)

    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    log(f"\n  Onglets ({len(sheet_names)}): {', '.join(sheet_names)}", output_lines)

    for sheet_name in sheet_names:
        log(f"\n  {'─' * 60}", output_lines)
        log(f"  ONGLET: {sheet_name}", output_lines)
        log(f"  {'─' * 60}", output_lines)

        # Lire les premières lignes pour détecter le vrai header
        df_raw = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, nrows=MAX_LIGNES, header=None, engine='openpyxl')
        log(f"\n  Premières lignes brutes (détection du header):", output_lines)
        for idx, row in df_raw.head(10).iterrows():
            vals = [str(v)[:30] for v in row.values[:8]]
            log(f"    Ligne {idx}: {vals}", output_lines)

        # Essayer de trouver le bon header (chercher une ligne avec des noms de colonnes)
        header_row = None
        for idx, row in df_raw.iterrows():
            non_null = row.notna().sum()
            has_text = any(isinstance(v, str) and len(v) > 2 for v in row.values)
            if non_null >= 3 and has_text:
                # Vérifier si ça ressemble à un header
                vals = [str(v).upper() for v in row.values if pd.notna(v)]
                if any(k in ' '.join(vals) for k in ['CODE', 'COMMUNE', 'DEP', 'REG', 'GEO', 'INSEE', 'NOM', 'LIBELLE']):
                    header_row = idx
                    break

        if header_row is not None:
            log(f"\n  Header détecté en ligne {header_row}", output_lines)
            df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, header=header_row, nrows=MAX_LIGNES, engine='openpyxl')
        else:
            log(f"\n  Header non détecté, utilisation de la ligne 0", output_lines)
            df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, nrows=MAX_LIGNES, engine='openpyxl')

        ws = wb[sheet_name]
        total_rows = ws.max_row - (header_row + 1 if header_row is not None else 1)
        log(f"  Lignes estimées: {total_rows:,}", output_lines)
        log(f"  Colonnes: {len(df.columns)}", output_lines)

        log(f"\n  COLONNES ({len(df.columns)}):", output_lines)
        for i, col in enumerate(df.columns, 1):
            dtype = str(df[col].dtype)
            non_null = df[col].notna().sum()
            log(f"    {i:3}. {str(col):<45} | type: {dtype:<10} | non-null: {non_null}/{len(df)}", output_lines)

        # Colonnes géographiques
        geo_cols = [c for c in df.columns if any(k in str(c).upper() for k in
            ['CODGEO', 'CODE', 'COMMUNE', 'DEP', 'REGION', 'GEO', 'INSEE'])]
        if geo_cols:
            log(f"\n  Colonnes géographiques: {geo_cols}", output_lines)

        # Aperçu
        log(f"\n  APERÇU (5 premières lignes):", output_lines)
        sample = df.head(5).to_string(max_colwidth=40)
        for line in sample.split('\n'):
            log(f"    {line}", output_lines)

        # Stats numériques
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        if numeric_cols:
            log(f"\n  STATS (colonnes numériques):", output_lines)
            stats = df[numeric_cols].describe().round(2)
            for line in stats.to_string().split('\n'):
                log(f"    {line}", output_lines)

        # Valeurs uniques catégorielles
        cat_cols = df.select_dtypes(include='object').columns.tolist()
        if cat_cols:
            log(f"\n  VALEURS UNIQUES (colonnes texte):", output_lines)
            for col in cat_cols:
                n_unique = df[col].nunique()
                if n_unique <= 20:
                    vals = df[col].value_counts().head(10)
                    log(f"    {col} ({n_unique} valeurs):", output_lines)
                    for val, count in vals.items():
                        log(f"      - {str(val)[:50]}: {count}", output_lines)
                else:
                    log(f"    {col}: {n_unique} valeurs uniques", output_lines)

    wb.close()

    log(f"\n{'=' * 70}", output_lines)
    log("FIN DE L'ANALYSE", output_lines)
    log("=" * 70, output_lines)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))

    print(f"\n=> Résultats sauvegardés dans: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
